from __future__ import annotations

import argparse
import csv
import io
import zipfile
from collections import defaultdict, deque
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
from openpyxl.styles import PatternFill

INVOICE_DELIMITER = ";"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_numeric(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).replace("\u00a0", " ").replace(" ", "").replace(",", ".").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def left_until_underscore(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    idx = text.find("_", 15)
    if idx == -1:
        return text
    return text[:idx]


def is_excel_file(path: Path) -> bool:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return True
    return zipfile.is_zipfile(path)


def read_order_rows(path: Path) -> list[list[object]]:
    if is_excel_file(path):
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            workbook = openpyxl.load_workbook(path, data_only=True)
        else:
            # The file is a zipped XLSX with a non-xlsx extension (e.g. .csv).
            data = path.read_bytes()
            workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.get_dialect("excel")
        reader = csv.reader(f, dialect)
        return [list(row) for row in reader]


def load_invoice_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    raw = path.read_bytes()
    encoding = "utf-8-sig"
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        encoding = "utf-16"
    text = raw.decode(encoding)
    reader = csv.DictReader(io.StringIO(text), delimiter=INVOICE_DELIMITER)
    return list(reader), reader.fieldnames or []


def build_invoice_index(
    rows: Iterable[dict[str, str]],
) -> dict[str, deque[dict[str, str]]]:
    index: dict[str, deque[dict[str, str]]] = defaultdict(deque)
    for row in rows:
        code = normalize_text(row.get("kod"))
        index[code].append(row)
    return index


def numbers_equal(left: object, right: object) -> bool:
    left_num = normalize_numeric(left)
    right_num = normalize_numeric(right)
    if left_num is None and right_num is None:
        return True
    if left_num is None or right_num is None:
        return False
    return left_num == right_num


def get_column(row: Sequence[object], index: int) -> object:
    if index - 1 < 0 or index - 1 >= len(row):
        return None
    return row[index - 1]


def compare_rows(
    order_rows: list[list[object]],
    invoice_rows: list[dict[str, str]],
    invoice_header: list[str],
) -> tuple[list[list[object]], list[bool]]:
    invoice_index = build_invoice_index(invoice_rows)

    output_rows: list[list[object]] = []
    row_matches: list[bool] = []

    if not order_rows:
        return output_rows, row_matches

    header = list(order_rows[0])
    header += ["processed_kod", "status", "mismatch_details"]
    output_rows.append(header)
    row_matches.append(True)

    for row in order_rows[1:]:
        order_kod_raw = get_column(row, 4)
        order_kod = left_until_underscore(order_kod_raw)
        order_qty = get_column(row, 6)
        order_unit = get_column(row, 9)
        order_net = get_column(row, 10)

        invoice_queue = invoice_index.get(order_kod)
        invoice_row = invoice_queue.popleft() if invoice_queue else None

        mismatches: list[str] = []
        if invoice_row is None:
            mismatches.append("missing_invoice_row")
        else:
            if normalize_text(order_kod) != normalize_text(invoice_row.get("kod")):
                mismatches.append("kod")
            if not numbers_equal(order_qty, invoice_row.get("db")):
                mismatches.append("db")
            if not numbers_equal(order_unit, invoice_row.get("egyseg_ar")):
                mismatches.append("egyseg_ar")
            if not numbers_equal(order_net, invoice_row.get("netto_ar")):
                mismatches.append("netto_ar")

        status = "OK" if not mismatches else "Mismatch"
        mismatch_details = ", ".join(mismatches)

        output_rows.append(list(row) + [order_kod, status, mismatch_details])
        row_matches.append(not mismatches)

    return output_rows, row_matches


def build_order_index(
    order_rows: list[list[object]],
) -> dict[str, deque[list[object]]]:
    index: dict[str, deque[list[object]]] = defaultdict(deque)
    for row in order_rows[1:]:
        raw = get_column(row, 4)
        code = left_until_underscore(raw)
        index[code].append(row)
    return index


def compare_invoice_rows(
    order_rows: list[list[object]],
    invoice_rows: list[dict[str, str]],
    invoice_header: list[str],
) -> tuple[list[list[object]], list[bool]]:
    order_index = build_order_index(order_rows)
    header = list(invoice_header) + ["status", "mismatch_details"]
    output_rows: list[list[object]] = [header]
    row_matches: list[bool] = [True]

    for invoice_row in invoice_rows:
        invoice_code = normalize_text(invoice_row.get("kod"))
        order_queue = order_index.get(invoice_code)
        order_row = order_queue.popleft() if order_queue else None

        mismatches: list[str] = []
        if order_row is None:
            mismatches.append("missing_order_row")
        else:
            order_qty = get_column(order_row, 6)
            order_unit = get_column(order_row, 9)
            order_net = get_column(order_row, 10)
            if not numbers_equal(order_qty, invoice_row.get("db")):
                mismatches.append("db")
            if not numbers_equal(order_unit, invoice_row.get("egyseg_ar")):
                mismatches.append("egyseg_ar")
            if not numbers_equal(order_net, invoice_row.get("netto_ar")):
                mismatches.append("netto_ar")

        status = "OK" if not mismatches else "Mismatch"
        mismatch_details = ", ".join(mismatches)
        row_values = [invoice_row.get(name, "") for name in invoice_header]
        output_rows.append(row_values + [status, mismatch_details])
        row_matches.append(not mismatches)

    return output_rows, row_matches


def write_report(
    rows: list[list[object]],
    row_matches: list[bool],
    sheet_name: str,
    workbook: openpyxl.Workbook | None = None,
) -> openpyxl.Workbook:
    if workbook is None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    for row in rows:
        sheet.append(row)

    max_col = sheet.max_column
    for idx, matches in enumerate(row_matches, start=1):
        if idx == 1:
            continue
        fill = GREEN_FILL if matches else RED_FILL
        for col in range(1, max_col + 1):
            sheet.cell(row=idx, column=col).fill = fill

    return workbook


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare an order Excel/CSV against invoice-output.csv."
    )
    parser.add_argument(
        "--order",
        type=Path,
        default=Path("order") / "belso megrendeles.csv",
        help="Path to the order Excel/CSV file.",
    )
    parser.add_argument(
        "--invoice",
        type=Path,
        default=Path("samples") / "invoice-output.csv",
        help="Path to the generated invoice-output.csv file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("order") / "compare-output.xlsx",
        help="Path to write the colored comparison report.",
    )
    parser.add_argument(
        "--base",
        choices=("order", "invoice"),
        default="order",
        help="Choose which side is the base of the comparison.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    order_path: Path = args.order
    invoice_path: Path = args.invoice
    output_path: Path = args.output

    if not order_path.is_file():
        parser.error(f"Order file not found: {order_path}")
    if not invoice_path.is_file():
        parser.error(f"Invoice file not found: {invoice_path}")

    order_rows = read_order_rows(order_path)
    invoice_rows, invoice_header = load_invoice_rows(invoice_path)

    if not order_rows:
        parser.error("No order rows found to compare.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    order_output_rows, order_row_matches = compare_rows(
        order_rows, invoice_rows, invoice_header
    )
    invoice_output_rows, invoice_row_matches = compare_invoice_rows(
        order_rows, invoice_rows, invoice_header
    )
    workbook = write_report(
        order_output_rows,
        order_row_matches,
        sheet_name="Order_to_Invoice",
    )
    workbook = write_report(
        invoice_output_rows,
        invoice_row_matches,
        sheet_name="Invoice_to_Order",
        workbook=workbook,
    )
    workbook.save(output_path)


if __name__ == "__main__":
    main()
