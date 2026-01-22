from __future__ import annotations

import csv
import io
import sys
import zipfile
from pathlib import Path
from typing import Sequence

from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from pypdf import PdfReader

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from main import (  # noqa: E402
    DEFAULT_TRANSLATIONS,
    FIELDNAMES,
    apply_translations,
    extract_page_text,
    format_currency,
    load_translations,
    parse_rows,
)
from order_compare import compare_invoice_rows, compare_rows, write_report  # noqa: E402


app = Flask(
    __name__,
    template_folder=str(PROJECT_ROOT / "templates"),
)


def read_order_rows_from_bytes(data: bytes) -> list[list[object]]:
    if zipfile.is_zipfile(io.BytesIO(data)):
        workbook = load_workbook(io.BytesIO(data), data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    encoding = "utf-8-sig"
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        encoding = "utf-16"
    text = data.decode(encoding)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.get_dialect("excel")
    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def build_invoice_rows(pdf_bytes: bytes) -> list[dict[str, str]]:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = extract_page_text(reader.pages)
    rows = parse_rows(text)
    translations = load_translations(DEFAULT_TRANSLATIONS)
    rows = apply_translations(rows, translations)
    for row in rows:
        row["egyseg_ar"] = format_currency(row["egyseg_ar"])
        row["netto_ar"] = format_currency(row["netto_ar"])
    return rows


def ensure_rows(rows: Sequence[Sequence[object]]) -> None:
    if not rows:
        raise ValueError("No rows found in the uploaded order file.")


@app.route("/", methods=["GET"])
def index() -> str:
    return render_template("index.html")


@app.route("/compare", methods=["POST"])
def compare() -> object:
    invoice_file = request.files.get("invoice_pdf")
    order_file = request.files.get("order_file")
    if not invoice_file or not order_file:
        return render_template(
            "index.html",
            error="Please upload both the invoice PDF and the order file.",
        )

    invoice_bytes = invoice_file.read()
    order_bytes = order_file.read()

    order_rows = read_order_rows_from_bytes(order_bytes)
    ensure_rows(order_rows)

    invoice_rows = build_invoice_rows(invoice_bytes)
    invoice_header = list(FIELDNAMES)

    order_output_rows, order_row_matches = compare_rows(
        order_rows, invoice_rows, invoice_header
    )
    invoice_output_rows, invoice_row_matches = compare_invoice_rows(
        order_rows, invoice_rows, invoice_header
    )

    workbook = write_report(
        order_output_rows,
        order_row_matches,
        sheet_name="Order_to_Invoice",
    )
    workbook = write_report(
        invoice_output_rows,
        invoice_row_matches,
        sheet_name="Invoice_to_Order",
        workbook=workbook,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="compare-output.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
