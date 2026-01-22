from __future__ import annotations

import argparse
import csv
import io
import zipfile
from pathlib import Path

import openpyxl


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def left_until_underscore(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    idx = text.find("_", 15)
    if idx == -1:
        return text
    return text[:idx]


def is_excel_file(path: Path) -> bool:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return True
    return zipfile.is_zipfile(path)


def read_order_rows(path: Path) -> list[list[object]]:
    if is_excel_file(path):
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            workbook = openpyxl.load_workbook(path, data_only=True)
        else:
            data = path.read_bytes()
            workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.get_dialect("excel")
        reader = csv.reader(f, dialect)
        return [list(row) for row in reader]


def write_output(path: Path, rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def build_output_rows(order_rows: list[list[object]]) -> list[list[object]]:
    output: list[list[object]] = []
    if not order_rows:
        return output

    header = ["alkatr_szam", "alkatr_szam_bal"]
    output.append(header)

    for row in order_rows[1:]:
        raw = row[3] if len(row) > 3 else None
        output.append([raw, left_until_underscore(raw)])

    return output


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract processed column 4 values from an order file."
    )
    parser.add_argument(
        "--order",
        type=Path,
        default=Path("order") / "belso-megrendeles.csv",
        help="Path to the order Excel/CSV file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("order") / "alkatr_szam_bal.xlsx",
        help="Path to write the extracted Excel file.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    order_path: Path = args.order
    output_path: Path = args.output

    if not order_path.is_file():
        parser.error(f"Order file not found: {order_path}")

    order_rows = read_order_rows(order_path)
    output_rows = build_output_rows(order_rows)
    if not output_rows:
        parser.error("No order rows found.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output(output_path, output_rows)


if __name__ == "__main__":
    main()
